import os
import faiss
import numpy as np
import pickle
import openpyxl
from sentence_transformers import SentenceTransformer
from django.db import connection
from search.Services.ElasticServices import create_document
from search.Services.normalizer import normalize
from search.models.postgres.models import LogUserSearch
from django.utils.timezone import now



class SbertService:
    def __init__(self, index_file='data/faiss_index.index', model_name='paraphrase-multilingual-mpnet-base-v2'):
    # def __init__(self, index_file='data/faiss_index.index', model_name='models/sbert-farsi-cosine/final'):
        self.model = SentenceTransformer(model_name, device="cpu")
        self.index_file = index_file
        self.index = None
        self.layers = []
        self.organizations = []

        directory = os.path.dirname(self.index_file)
        if not os.path.exists(directory):
            os.makedirs(directory)

        if os.path.exists(self.index_file):
            self.index = faiss.read_index(self.index_file)
        else:
            self.index = None

    def normalize_embedding(self, embedding):
        norm = np.linalg.norm(embedding)
        if norm != 0:
            embedding = embedding / norm
        return embedding

    def encode_and_store_embedding(self, layer, system, data_model):
        normalize_layer = normalize(layer)
        embedding = self.model.encode([normalize_layer], device="cpu")[0].astype(np.float32)
        normalized_embedding = self.normalize_embedding(embedding)
        binary_embedding = pickle.dumps(normalized_embedding)

        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id FROM embeddings 
                WHERE layer = %s AND system = %s AND "dataModel" = %s
            """, (layer, system, data_model))
            existing = cursor.fetchone()

            if existing:
                stored_id = existing[0]
            else:
                cursor.execute("""
                    INSERT INTO embeddings ("layer", "system", "dataModel", "embeddingLayer")
                    VALUES (%s, %s, %s, %s)
                    RETURNING id;
                """, (layer, system, data_model, binary_embedding))
                row = cursor.fetchone()
                connection.commit()
                stored_id = row[0] if row else None

                if stored_id is None:
                    print("Failed to retrieve stored ID.")
                    return (None, False)


                # data = {"layer": layer, "system": system, "dataModel": data_model}
                # response_elastic = create_document(data)
                # print(response_elastic)


        if self.index is None:
            flat_index = faiss.IndexFlatL2(normalized_embedding.shape[0])
            self.index = faiss.IndexIDMap(flat_index)
            print("Faiss index created with IDMap.")

        if self.index is not None:
            _, I = self.index.search(np.array([normalized_embedding]), k=100)
            if stored_id in I[0]:
                print(f"ID {stored_id} already exists in FAISS index. Skipping add.")
                return (stored_id, True)
            else:
                self.index.add_with_ids(np.array([normalized_embedding]), np.array([stored_id], dtype=np.int64))
        else:
            self.index.add_with_ids(np.array([normalized_embedding]), np.array([stored_id], dtype=np.int64))

        faiss.write_index(self.index, self.index_file)

        print(f"Stored ID: {stored_id}")
        print(f"Total embeddings in Faiss index: {self.index.ntotal}")
        return (stored_id, existing is not None)

    def process_excel_file(self, excel_file, min_row=2, max_row=None):
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        messages = []

        if max_row is None:
            max_row = sheet.max_row

        for idx, row in enumerate(sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True), start=min_row):
            if not row:
                continue

            if len(row) < 3:
                messages.append(f"Row {idx} has incomplete data: {row}")
                continue

            layer, system, data_model = row[:3]
            # layer = normalize(layer)
            system = normalize(system)

            try:
                data_model = int(data_model)
            except (ValueError, TypeError):
                messages.append(f"Row {idx} has invalid dataModel: {data_model}")
                continue

            self.encode_and_store_embedding(layer, system, data_model)
            # stored_id, is_duplicate = self.encode_and_store_embedding(layer, system, data_model)
            # if is_duplicate:
            #     messages.append(f"Row {idx}: Record already exists for layer: {layer}, system: {system}, dataModel: {data_model}")
            # else:
            #     messages.append(f"Row {idx} processed. Stored ID: {stored_id}")
        if messages is not None and len(messages) < 1:
            messages.append("عملیات با موفقیت انجام شد")
        return messages

    def search(self, query, top_k=5):
        if self.index is None:
            raise ValueError("Faiss index has not been created. Ensure 'encode_and_store_embedding' is called first.")
        query = normalize(query)
        query_embedding = self.model.encode([query], device="cpu").astype(np.float32)

        D, I = self.index.search(query_embedding, top_k)
        valid_ids = [int(idx) for idx in I[0] if idx != -1]

        if not valid_ids:
            return []

        placeholders = ', '.join(['%s'] * len(valid_ids))
        sql = f"""SELECT id, layer, system, "dataModel"
                  FROM embeddings
                  WHERE id IN ({placeholders})"""
        with connection.cursor() as cursor:
            cursor.execute(sql, valid_ids)
            rows = cursor.fetchall()

        rows_dict = {row[0]: (row[1], row[2], row[3]) for row in rows}
        results = []
        for idx in valid_ids:
            if idx == -1 or idx not in rows_dict:
                continue
            layer, system, data_model = rows_dict[idx]
            distance = D[0][list(I[0]).index(idx)]
            normalized_score = 1 / (1 + distance)
            results.append({
                "layer": layer,
                "system": system,
                "score": normalized_score,
                "dataModel": data_model
            })
        return sorted(results, key=lambda x: x['score'], reverse=True)

def save_log_entry_from_json_sbert(data):

    if isinstance(data, list):
        entries = data
    else:
        entries = [data]

    for item in entries:
        query = item.get('query')
        layer = item.get('layer')
        organization = item.get('organization')
        relevance = int(item.get('relevance', 0))
        score = float(item.get('score', 0))
        type_result = int(item.get('type_result', 1))

        log_entry = LogUserSearch(
            query=query,
            selected_layer=layer,
            selected_organization=organization,
            relevance=relevance,
            score=score,
            type_result=type_result,
            timestamp=now()
        )
        log_entry.save()


sbert_service = SbertService(index_file='data/faiss_index.index')