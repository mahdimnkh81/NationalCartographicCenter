import faiss
import pandas as pd
from datasets import Dataset
from search.models.postgres.models import LogUserSearch
from collections import defaultdict
import numpy as np
from sentence_transformers import SentenceTransformer, losses, util
from search.Services.normalizer import normalize
from django.db import connection
import openpyxl
import pickle
import os
import shutil
from transformers import AutoModel
from sentence_transformers import (
    SentenceTransformer,
    SentenceTransformerTrainer,
    SentenceTransformerTrainingArguments,
    SentenceTransformerModelCardData,
    losses,
    evaluation
)
from sentence_transformers.evaluation import EmbeddingSimilarityEvaluator

class TrainService:
    # def __init__(self, index_file='data/trainModel_faiss_index.index', model_name='models/sbert-farsi-cosine/final'):
    def __init__(self, index_file='data/faiss_index.index', model_name='paraphrase-multilingual-mpnet-base-v2'):
        self.model = SentenceTransformer(model_name, device="cpu")
        self.index_file = index_file
        self.index = None
        self.layers = []
        self.organizations = []

        directory = os.path.dirname(self.index_file)
        if not os.path.exists(directory):
            os.makedirs(directory)

        if os.path.exists(self.index_file):
            self.index = faiss.read_index(self.index_file)
        else:
            self.index = None

    def normalize_embedding(self, embedding):
        norm = np.linalg.norm(embedding)
        if norm != 0:
            embedding = embedding / norm
        return embedding

    def encode_and_store_embedding(self, layer, system, data_model):
        normalize_layer = normalize(layer)
        embedding = self.model.encode([normalize_layer], device="cpu")[0].astype(np.float32)
        normalized_embedding = self.normalize_embedding(embedding)
        binary_embedding = pickle.dumps(normalized_embedding)

        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id FROM embeddings_train_model
                WHERE layer = %s AND system = %s AND "dataModel" = %s
            """, (layer, system, data_model))
            existing = cursor.fetchone()

            if existing:
                stored_id = existing[0]
            else:
                cursor.execute("""
                    INSERT INTO embeddings_train_model ("layer", "system", "dataModel")
                    VALUES (%s, %s, %s)
                    RETURNING id;
                """, (layer, system, data_model))
                row = cursor.fetchone()
                connection.commit()
                stored_id = row[0] if row else None

                if stored_id is None:
                    print("Failed to retrieve stored ID.")
                    return (None, False)


                # data = {"layer": layer, "system": system, "dataModel": data_model}
                # response_elastic = create_document(data)
                # print(response_elastic)


        if self.index is None:
            flat_index = faiss.IndexFlatL2(normalized_embedding.shape[0])
            self.index = faiss.IndexIDMap(flat_index)
            print("Faiss index created with IDMap.")

        if self.index is not None:
            _, I = self.index.search(np.array([normalized_embedding]), k=100)
            if stored_id in I[0]:
                print(f"ID {stored_id} already exists in FAISS index. Skipping add.")
                return (stored_id, True)
            else:
                self.index.add_with_ids(np.array([normalized_embedding]), np.array([stored_id], dtype=np.int64))
        else:
            self.index.add_with_ids(np.array([normalized_embedding]), np.array([stored_id], dtype=np.int64))

        faiss.write_index(self.index, self.index_file)

        print(f"Stored ID: {stored_id}")
        print(f"Total embeddings in Faiss index: {self.index.ntotal}")
        return (stored_id, existing is not None)

    def process_excel_file(self, excel_file, min_row=2, max_row=None):
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        messages = []

        if max_row is None:
            max_row = sheet.max_row

        for idx, row in enumerate(sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True), start=min_row):
            if not row:
                continue

            if len(row) < 3:
                messages.append(f"Row {idx} has incomplete data: {row}")
                continue

            layer, system, data_model = row[:3]
            # layer = normalize(layer)
            system = normalize(system)

            try:
                data_model = int(data_model)
            except (ValueError, TypeError):
                messages.append(f"Row {idx} has invalid dataModel: {data_model}")
                continue

            self.encode_and_store_embedding(layer, system, data_model)
            # stored_id, is_duplicate = self.encode_and_store_embedding(layer, system, data_model)
            # if is_duplicate:
            #     messages.append(f"Row {idx}: Record already exists for layer: {layer}, system: {system}, dataModel: {data_model}")
            # else:
            #     messages.append(f"Row {idx} processed. Stored ID: {stored_id}")
        if messages is not None and len(messages) < 1:
            messages.append("عملیات با موفقیت انجام شد")
        return messages

    def search(self, query, top_k=5):
        if self.index is None:
            raise ValueError("Faiss index has not been created. Ensure 'encode_and_store_embedding' is called first.")
        query = normalize(query)
        query_embedding = self.model.encode([query], device="cpu").astype(np.float32)

        D, I = self.index.search(query_embedding, top_k)
        valid_ids = [int(idx) for idx in I[0] if idx != -1]

        if not valid_ids:
            return []

        placeholders = ', '.join(['%s'] * len(valid_ids))
        sql = f"""SELECT id, layer, system, "dataModel"
                  FROM embeddings_train_model
                  WHERE id IN ({placeholders})"""
        with connection.cursor() as cursor:
            cursor.execute(sql, valid_ids)
            rows = cursor.fetchall()

        rows_dict = {row[0]: (row[1], row[2], row[3]) for row in rows}
        results = []
        for idx in valid_ids:
            if idx == -1 or idx not in rows_dict:
                continue
            layer, system, data_model = rows_dict[idx]
            distance = D[0][list(I[0]).index(idx)]
            normalized_score = 1 / (1 + distance)
            results.append({
                "layer": layer,
                "system": system,
                "score": normalized_score,
                "dataModel": data_model
            })
        return sorted(results, key=lambda x: x['score'], reverse=True)


train_service = TrainService(index_file='data/trainModel_faiss_index.index')



















def build_sts_from_db(
    include_unlabeled: bool = False,  # اگر False باشد، نمونه‌های relevance=0 حذف می‌شوند
    min_len: int = 1,
    seed: int = 42
):
    # qs = LogUserSearch.objects.values('query', 'selected_layer', 'relevance')
    qs = (
        LogUserSearch.objects
        .values('query', 'selected_layer', 'relevance')
        .order_by('-id')[:2000]
    )
    rows = []
    for r in qs:
        s1 = (r.get('query') or '').strip()
        s2 = (r.get('selected_layer') or '').strip()
        rel = r.get('relevance')
        if not s1 or not s2:
            continue
        if rel is None:
            continue

        # نگاشت هدف فقط از روی relevance
        if rel == 1:
            y = 1.0
        elif rel == -1:
            y = 0.0
        elif rel == 0:
            if not include_unlabeled:
                continue  # نمونه‌های بدون فیدبک را حذف کن
            y = 0.5
        else:
            # اگر مقدار غیرمنتظره است، ردش کن
            continue

        rows.append({"sentence1": s1, "sentence2": s2, "score": y})

    df = pd.DataFrame(rows)
    if df.empty:
        raise ValueError("هیچ دادهٔ معتبری برای آموزش پیدا نشد.")

    # تمیزکاری سبک
    # df = df[(df['sentence1'].str.len() >= min_len) & (df['sentence2'].str.len() >= min_len)]
    # df = df.drop_duplicates(subset=['sentence1', 'sentence2'])
    # if df.empty:
    #     raise ValueError("پس از تمیزکاری، دیتایی باقی نماند.")

    # اطمینان از نوع داده‌ها
    df['score'] = df['score'].astype(float)  # تبدیل امتیازها به float
    df['sentence1'] = df['sentence1'].astype(str)  # تبدیل جملات به string
    df['sentence2'] = df['sentence2'].astype(str)  # تبدیل جملات به string

    # ساخت دیتاست HuggingFace
    dataset = Dataset.from_pandas(df[['sentence1', 'sentence2', 'score']], preserve_index=False)

    # تقسیم 80/10/10
    split = dataset.train_test_split(test_size=0.1, seed=seed, shuffle=True)  # 10% → test
    val_from_train = 0.1 / 0.9
    train_eval = split['train'].train_test_split(test_size=val_from_train, seed=seed, shuffle=True)

    return train_eval['train'], train_eval['test'], split['test']

def evaluate_ir(model, dataset, ks=[1, 2, 5]):
    """
    محاسبه P@k، Success@k، MRR و MAP برای چند k مختلف.
    Success@k یعنی آیا حداقل یک سند مرتبط در بین top-k هست یا نه.
    """
    # 1. تبدیل دیتاست به pandas
    df = dataset.to_pandas() if not isinstance(dataset, pd.DataFrame) else dataset


    # 3. گروه‌بندی بر اساس query
    queries = defaultdict(list)
    for _, row in df.iterrows():
        queries[row["sentence1"]].append((row["sentence2"], float(row["score"])))


    # 4. محاسبه embedding ها
    all_sentences = list(set(df["sentence1"].tolist() + df["sentence2"].tolist()))
    embeddings = model.encode(all_sentences, convert_to_tensor=True, show_progress_bar=False)
    emb_map = {s: embeddings[i] for i, s in enumerate(all_sentences)}

    # 5. متریک‌ها
    precisions_at_k = {k: [] for k in ks}
    success_at_k = {k: [] for k in ks}
    reciprocal_ranks, average_precisions = [], []

    for q, docs in queries.items():
        q_emb = emb_map[q]
        doc_embs = [emb_map[d] for d, _ in docs]

        # similarity
        sims = np.array([float(util.cos_sim(q_emb, d_emb)) for d_emb in doc_embs])

        # sort by similarity
        ranked = sorted(zip(docs, sims), key=lambda x: x[1], reverse=True)

        # relevance labels
        labels = [int(doc[1] > 0.5) for doc, _ in ranked]  # >0.5 = relevant

        # P@k و Success@k برای همه‌ی ks
        for k in ks:
            if len(labels) >= k:
                topk = labels[:k]
            else:
                topk = labels

            precisions_at_k[k].append(sum(topk) / k)
            success_at_k[k].append(1 if any(topk) else 0)  # ✅ اگر حداقل یک مرتبط باشد = موفقیت

        # MRR
        rr = 0
        for rank, rel in enumerate(labels, start=1):
            if rel == 1:
                rr = 1.0 / rank
                break
        reciprocal_ranks.append(rr)

        # AP (برای MAP)
        num_rel, prec_sum = 0, 0.0
        for rank, rel in enumerate(labels, start=1):
            if rel == 1:
                num_rel += 1
                prec_sum += num_rel / rank
        ap = prec_sum / num_rel if num_rel > 0 else 0
        average_precisions.append(ap)

    # ساخت دیکشنری خروجی
    metrics = {}

    # میانگین P@k و Success@k
    for k in ks:
        metrics[f"P@{k}"] = float(np.mean(precisions_at_k[k]))
        metrics[f"Success@{k}"] = float(np.mean(success_at_k[k]))

    # سایر معیارها
    metrics.update({
        "MRR": float(np.mean(reciprocal_ranks)),
        "MAP": float(np.mean(average_precisions))
    })

    return metrics


def train_sbert_model(SEED=42, num_train_epochs=1, batch_size=8, learning_rate=2e-5):
    """آموزش مدل و بازگشت نتیجه تست قبل و بعد آموزش به صورت dict JSON-safe."""
    np.random.seed(SEED)

    # 1) load base model
    model = SentenceTransformer(
        "sentence-transformers/paraphrase-mpnet-base-v2",
        model_card_data=SentenceTransformerModelCardData(
            language="fa",
            license="apache-2.0",
            model_name="SBERT Farsi Cosine (relevance-supervised)"
        )
    )

    # 2) build datasets
    train_dataset, eval_dataset, test_dataset = build_sts_from_db(
        include_unlabeled=False, min_len=2, seed=SEED
    )

    # 3) loss
    train_loss = losses.CosineSimilarityLoss(model=model)

    # 4) حذف مدل قبلی
    models_dir = "models/sbert-farsi-cosine"
    if os.path.exists(models_dir):
        shutil.rmtree(models_dir)

    # 5) training args
    args = SentenceTransformerTrainingArguments(
        output_dir=models_dir,
        num_train_epochs=num_train_epochs,
        per_device_train_batch_size=batch_size,
        per_device_eval_batch_size=batch_size,
        learning_rate=learning_rate,
        eval_strategy="no",
        save_strategy="steps",
        save_steps=200,
        save_total_limit=2,
        logging_steps=50,
        run_name="sbert-farsi-cosine",
        report_to=[],
        load_best_model_at_end=False,
    )

    # 6) trainer
    trainer = SentenceTransformerTrainer(
        model=model,
        args=args,
        train_dataset=train_dataset,
        eval_dataset=eval_dataset,
        loss=train_loss,
    )

    # 🔹 Pre-training evaluation
    pre_metrics = evaluate_ir(model, test_dataset, ks=[1,2,5])

    # 7) train
    trainer.train()

    # 8) save **کل مدل** با متد save خود SentenceTransformer
    model.save("models/sbert-farsi-cosine/final")  # ✅ این مهمه

    # 🔹 Post-training evaluation
    post_metrics = evaluate_ir(model, test_dataset, ks=[1,2,5])

    return {
        "train_size": len(train_dataset),
        "eval_size": len(eval_dataset),
        "test_size": len(test_dataset),
        "pre_train_metrics": pre_metrics,
        "post_train_metrics": post_metrics,
        "final_dir": "models/sbert-farsi-cosine/final"
    }




